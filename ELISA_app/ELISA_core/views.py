from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import Plates
import openpyxl
import json

# Create your views here.
def Home(request):
    return render(request, 'Home.html')

def Input_data(request):
    try:
        if request.method == 'POST':
            if request.POST.get('text_submit'):
                data = request.POST.get('text_input')
                if data == '':
                    return render(request, 'Input_data.html', {
                        'check': 'text',
                    })
                data = data.strip().split('\r\n')
                data_string = formatting_txt(data, 1)
                database(data_string)
            if request.POST.get('file_submit'):
                if request.FILES.getlist('my_file') == []:
                    return render(request, 'Input_data.html', {
                        'check': 'file',
                    })
                for file in request.FILES.getlist('my_file'):
                    if str(file).split('.')[1] not in ['txt', 'xlsx']:
                        return render(request, 'Input_data.html', {
                            'check': 'extension',
                        })
                for file in request.FILES.getlist('my_file'):
                    if str(file).split('.')[1] == 'txt':
                        data = file.readlines()
                        data_string = formatting_txt(data, 2)
                        database(data_string)
                    elif str(file).split('.')[1] == 'xlsx':
                        data_string = formatting_xlsx(file)
                        database(data_string)
            return render(request, 'Input_data.html', {
                'check': 'correct',
            })
        else:
            return render(request, 'Input_data.html')
    except:
        return render(request, 'Input_data.html', {
            'check': 'false',
        })

def formatting_txt(data, counter):
    lines = list()
    for i in data[:-1]:
        if counter == 1:
            lines.append(i.strip())
        elif counter == 2:
            lines.append(i.strip().decode('utf-8'))
    formatted_data = list()
    for j in lines:
        line = j.split('\t')
        formatted_data.append(line)
    formatted_data[1].insert(0, '#')
    data_string = ""
    for i in formatted_data:
        for j in i:
            data_string += j + "="
    return data_string

def formatting_xlsx(file_name):
    wb = openpyxl.load_workbook(file_name)
    active_sheet = wb.active
    excel_data = list()
    for row in active_sheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    formatted_data = excel_data
    del formatted_data[1][0]
    del formatted_data[0][1:]
    formatted_data[1].insert(0, '#')
    data_string = ""
    for i in formatted_data:
        for j in i:
            data_string += j + "="
    return data_string

id = 1

def database(data_string):
    global id
    split = data_string.split('=')
    plates_instance = Plates.objects.create(
        id=str(id),
        name=str(split[0]),
        data=data_string
    )
    id += 1

totaal = []
check = ''

def Plate_layout(request):
    global check
    if request.method == 'POST':
        if request.POST.get('file_submit'):
            if request.FILES.getlist("my_file") == []:
                check = 'error'
                return render(request, 'Plate_layout.html', {
                    'check': check,
                })
            excel_file = request.FILES["my_file"]
            wb = openpyxl.load_workbook(excel_file)
            active_sheet = wb.active
            excel_data = list()
            for row in active_sheet.iter_rows():
                row_data = list()
                for cell in row:
                    if type(cell.value) == float:
                        row_data.append(str(round(cell.value)))
                    else:
                        row_data.append(str(cell.value))
                excel_data.append(row_data)
            temp = []
            global totaal
            counter = 0
            for i in excel_data:
                i = [e for e in i if e not in ('None')]
                if len(i) != 0:
                    if counter == 1:
                        i.insert(0, '#')
                    temp.append(i)
                    counter += 1
                    if counter == 10:
                        totaal.append(temp)
                        counter = 0
                        temp = []
            check = 'go'
            return render(request, 'Plate_layout.html', {
                'totaal': totaal,
                'check': check,
            })
        if request.POST.get('standaard_input'):
            values = request.POST.get('standaard')
            counter = 0
            counter2 = 0
            for i in totaal:
                for j in i[2:]:
                    if counter2 != 7:
                        j[1] = float(values)
                        j[2] = float(values)
                        values = float(values)/2
                    elif counter2 == 7:
                        j[1] = '#'
                        j[2] = '#'
                    counter2 += 1
                counter += 1
                values = request.POST.get('standaard')
                counter2 = 0
            check = 'go'
            return render(request, 'Plate_layout.html', {
            'totaal': totaal,
            'check': check,
        })
    else:
        return render(request, 'Plate_layout.html', {
            'totaal': totaal,
            'check': check,
        })

end_dilution = []
def Dilutions(request):
    global end_dilution
    if request.method == 'POST':
        if request.POST.get('dilution_submit'):
            dilution = request.POST.get('dilution')
            ll = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
            end_list = [["#", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]]
            for i in range(9):
                temp = []
                for x in range(13):
                    if i == 0:
                        if x == 0:
                            temp.append(ll[i])
                        elif x == 1 or x == 2 or x == 3 or x == 8:
                            temp.append("1")
                        else:
                            temp.append(dilution)
                    else:
                        if x == 0:
                            temp.append(ll[i])
                        elif x == 1 or x == 2:
                            temp.append("1")
                        else:
                            temp.append(dilution)
                end_list.append(temp)
            end_dilution = end_list
            return render(request, 'Dilutions.html', {
                "end_list": end_list
            })
    return render(request, 'Dilutions.html', {
        "end_list": end_dilution})

def Visualize_data(request):
    data = Plates.objects.values()
    dictionary = {}
    teller = 1
    counter = 0
    nested = []
    temp = []
    for i in data:
        lines = i['data'].split('=')[:-1]
        print((float(lines[106]) + float(lines[107]))/2)
        for j in lines[1:]:
            temp.append(j)
            counter += 1
            if counter == 13:
                nested.append(temp)
                counter = 0
                temp = []
        dictionary[teller] = nested
        nested = []
        teller += 1
    return render(request, 'Visualize_data.html', {
        'dictionary': dictionary,
    })

def Cut_off(request):
    return render(request, 'Cut_off.html')

def Intermediate_result(request):
    return render(request, 'Intermediate_result.html')

def End_results(request):
    return render(request, 'End_results.html')
